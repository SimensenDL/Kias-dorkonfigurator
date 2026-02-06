"""
Excel-eksport for kappelister.

Eksporterer produksjonslisten til et formatert Excel-dokument
med gruppering per karmtype, som originale kappelister.
"""
from pathlib import Path
from typing import TYPE_CHECKING
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if TYPE_CHECKING:
    from ..models.production_list import ProductionList


def export_kappliste_excel(prod_list: 'ProductionList', filepath: str | Path) -> None:
    """Eksporterer kappeliste til Excel.

    Args:
        prod_list: ProductionList med dører
        filepath: Filsti for Excel-fil

    Raises:
        ImportError: Hvis openpyxl ikke er installert
        Exception: Ved eksportfeil
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl er ikke installert. Kjør: uv add openpyxl")

    filepath = Path(filepath)
    if not filepath.suffix:
        filepath = filepath.with_suffix('.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Kappeliste"

    # Stiler
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    subsection_font = Font(bold=True, size=11)
    subsection_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Tittel
    row = 1
    ws.cell(row=row, column=1, value="KAPPELISTE")
    ws.cell(row=row, column=1).font = Font(bold=True, size=18)
    row += 1

    # Dato og antall dører
    ws.cell(row=row, column=1, value=f"Generert: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Antall dører: {prod_list.door_count}")
    row += 2

    # Hent grupperte items
    grouped = prod_list.get_grouped_items()

    # Definer komponent-kategorier
    karm_komponenter = ['Overligger', 'Hengselside', 'Sluttstykkeside', 'Sidedel']
    dorblad_komponenter = ['Dørblad', 'Laminat', 'Dekklist']
    tilbehor_komponenter = ['Terskel', 'Løfteterskel', 'Skinne', 'Innerskinne', 'Styreskinne',
                            'Sparkeplate', 'Ryggforsterkning', 'Ryggforst. overdel', 'Avviserbøyler']

    # Iterér gjennom karmtyper
    for karm_type, komponenter in sorted(grouped.items()):
        # Karmtype-overskrift
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"Karmtype: {karm_type}")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_align
        row += 1

        # Karm-komponenter
        karm_items = {k: v for k, v in komponenter.items() if k in karm_komponenter}
        if karm_items:
            row = _write_category_section(ws, row, "Karm", karm_items,
                                          ['Komponent', 'Antall', 'Lengde', 'Side', 'Farge'],
                                          subsection_font, subsection_fill, thin_border,
                                          center_align, right_align, left_align,
                                          show_side=True)

        # Dørblad-komponenter
        dorblad_items = {k: v for k, v in komponenter.items() if k in dorblad_komponenter}
        if dorblad_items:
            row = _write_category_section(ws, row, "Dørblad", dorblad_items,
                                          ['Komponent', 'Antall', 'Bredde', 'Høyde', 'Farge'],
                                          subsection_font, subsection_fill, thin_border,
                                          center_align, right_align, left_align,
                                          show_dimensions=True)

        # Tilbehør-komponenter
        tilbehor_items = {k: v for k, v in komponenter.items() if k in tilbehor_komponenter}
        if tilbehor_items:
            row = _write_category_section(ws, row, "Tilbehør", tilbehor_items,
                                          ['Komponent', 'Antall', 'Lengde', 'Farge', ''],
                                          subsection_font, subsection_fill, thin_border,
                                          center_align, right_align, left_align)

        # Tom rad mellom karmtyper
        row += 1

    # Juster kolonnebredder
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Lagre
    wb.save(filepath)


def _write_category_section(ws, start_row: int, category: str, komponenter: dict,
                            headers: list, subsection_font, subsection_fill, thin_border,
                            center_align, right_align, left_align,
                            show_side: bool = False, show_dimensions: bool = False) -> int:
    """Skriver en kategoriseksjon til arket.

    Returns:
        Neste ledige rad
    """
    row = start_row

    # Kategori-overskrift
    cell = ws.cell(row=row, column=1, value=category)
    cell.font = subsection_font
    cell.fill = subsection_fill
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = subsection_fill
    row += 1

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = center_align
    row += 1

    # Data
    for komponent_navn, items in komponenter.items():
        for item in items:
            # Komponent
            cell = ws.cell(row=row, column=1, value=komponent_navn)
            cell.border = thin_border
            cell.alignment = left_align

            # Antall
            cell = ws.cell(row=row, column=2, value=item.antall)
            cell.border = thin_border
            cell.alignment = center_align

            if show_side:
                # Lengde
                cell = ws.cell(row=row, column=3, value=item.lengde or '-')
                cell.border = thin_border
                cell.alignment = right_align

                # Side
                cell = ws.cell(row=row, column=4, value=item.side or '-')
                cell.border = thin_border
                cell.alignment = center_align

                # Farge
                cell = ws.cell(row=row, column=5, value=item.farge or '-')
                cell.border = thin_border
                cell.alignment = center_align

            elif show_dimensions:
                # Bredde
                cell = ws.cell(row=row, column=3, value=item.bredde or '-')
                cell.border = thin_border
                cell.alignment = right_align

                # Høyde
                cell = ws.cell(row=row, column=4, value=item.hoyde or '-')
                cell.border = thin_border
                cell.alignment = right_align

                # Farge
                cell = ws.cell(row=row, column=5, value=item.farge or '-')
                cell.border = thin_border
                cell.alignment = center_align

            else:
                # Lengde
                cell = ws.cell(row=row, column=3, value=item.lengde or '-')
                cell.border = thin_border
                cell.alignment = right_align

                # Farge
                cell = ws.cell(row=row, column=4, value=item.farge or '-')
                cell.border = thin_border
                cell.alignment = center_align

            row += 1

    return row


def is_excel_available() -> bool:
    """Sjekker om Excel-eksport er tilgjengelig.

    Returns:
        True hvis openpyxl er installert
    """
    return HAS_OPENPYXL
